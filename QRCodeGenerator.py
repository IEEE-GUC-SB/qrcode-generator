# Importing libraries
import qrcode
import openpyxl
import sys

# Ensure correct usage
if len(sys.argv) != 2:
    sys.exit("Usage: python3 QRCodeGenerator.py <Excel file>")
fileName = sys.argv[1]

 # Define variable to load the dataframe
dataframe = openpyxl.load_workbook(fileName)
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(1, dataframe1.max_row):
    # Data to be encoded
    data = ''
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        data += str(col[row].value) +' '
    # Encoding data
    img = qrcode.make(data)
    # Saving as an image file
    img.save('MyQRCode' + str(row) +'.png')